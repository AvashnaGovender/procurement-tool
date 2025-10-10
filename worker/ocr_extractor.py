"""OCR and document extraction capabilities."""
import os
import io
from typing import Dict, List, Any, Optional
from pathlib import Path
import pytesseract
from PIL import Image
import PyPDF2
from docx import Document
import openpyxl
from pydantic import BaseModel


class ExtractedField(BaseModel):
    """Model for extracted field data."""
    field_name: str
    value: str
    confidence: float
    source_page: Optional[int] = None


class DocumentExtractionResult(BaseModel):
    """Model for document extraction results."""
    document_type: str
    extracted_text: str
    extracted_fields: List[ExtractedField]
    confidence_score: float
    processing_time: float


class OCRExtractor:
    """OCR and document extraction service."""
    
    def __init__(self):
        """Initialize OCR extractor."""
        # Configure tesseract path if needed
        # pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        pass
    
    def extract_from_image(self, image_path: str) -> str:
        """Extract text from image using OCR."""
        try:
            image = Image.open(image_path)
            text = pytesseract.image_to_string(image)
            return text.strip()
        except Exception as e:
            print(f"Error extracting text from image {image_path}: {e}")
            return ""
    
    def extract_from_pdf(self, pdf_path: str) -> Dict[str, Any]:
        """Extract text and metadata from PDF."""
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                page_count = len(pdf_reader.pages)
                
                for page_num in range(page_count):
                    page = pdf_reader.pages[page_num]
                    text += page.extract_text() + "\n"
                
                return {
                    "text": text.strip(),
                    "page_count": page_count,
                    "metadata": pdf_reader.metadata
                }
        except Exception as e:
            print(f"Error extracting text from PDF {pdf_path}: {e}")
            return {"text": "", "page_count": 0, "metadata": {}}
    
    def extract_from_docx(self, docx_path: str) -> str:
        """Extract text from DOCX file."""
        try:
            doc = Document(docx_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            print(f"Error extracting text from DOCX {docx_path}: {e}")
            return ""
    
    def extract_from_excel(self, excel_path: str) -> Dict[str, Any]:
        """Extract data from Excel file."""
        try:
            workbook = openpyxl.load_workbook(excel_path)
            data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        sheet_data.append(list(row))
                
                data[sheet_name] = sheet_data
            
            return data
        except Exception as e:
            print(f"Error extracting data from Excel {excel_path}: {e}")
            return {}
    
    def extract_fields_from_text(self, text: str, document_type: str) -> List[ExtractedField]:
        """Extract specific fields from document text."""
        fields = []
        
        # Common field patterns for supplier documents
        field_patterns = {
            "company_name": [
                r"Company\s+Name[:\s]+([^\n]+)",
                r"Business\s+Name[:\s]+([^\n]+)",
                r"Trading\s+Name[:\s]+([^\n]+)"
            ],
            "registration_number": [
                r"Registration\s+Number[:\s]+([^\n]+)",
                r"Company\s+Registration[:\s]+([^\n]+)",
                r"Reg\s+No[:\s]+([^\n]+)"
            ],
            "vat_number": [
                r"VAT\s+Number[:\s]+([^\n]+)",
                r"VAT\s+Reg[:\s]+([^\n]+)"
            ],
            "contact_person": [
                r"Contact\s+Person[:\s]+([^\n]+)",
                r"Responsible\s+Person[:\s]+([^\n]+)"
            ],
            "phone": [
                r"Phone[:\s]+([^\n]+)",
                r"Telephone[:\s]+([^\n]+)",
                r"Contact\s+Number[:\s]+([^\n]+)"
            ],
            "email": [
                r"Email[:\s]+([^\n]+)",
                r"E-mail[:\s]+([^\n]+)"
            ],
            "address": [
                r"Address[:\s]+([^\n]+)",
                r"Physical\s+Address[:\s]+([^\n]+)"
            ]
        }
        
        import re
        
        for field_name, patterns in field_patterns.items():
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    fields.append(ExtractedField(
                        field_name=field_name,
                        value=match.group(1).strip(),
                        confidence=0.8  # Basic confidence score
                    ))
                    break
        
        return fields
    
    def process_document(self, file_path: str, document_type: str) -> DocumentExtractionResult:
        """Process a document and extract text and fields."""
        import time
        start_time = time.time()
        
        file_extension = Path(file_path).suffix.lower()
        extracted_text = ""
        extracted_fields = []
        
        try:
            if file_extension in ['.pdf']:
                pdf_result = self.extract_from_pdf(file_path)
                extracted_text = pdf_result["text"]
            elif file_extension in ['.docx']:
                extracted_text = self.extract_from_docx(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                excel_data = self.extract_from_excel(file_path)
                # Convert Excel data to text for field extraction
                extracted_text = str(excel_data)
            elif file_extension in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
                extracted_text = self.extract_from_image(file_path)
            else:
                # Try OCR for unknown formats
                extracted_text = self.extract_from_image(file_path)
            
            # Extract specific fields
            extracted_fields = self.extract_fields_from_text(extracted_text, document_type)
            
            processing_time = time.time() - start_time
            
            return DocumentExtractionResult(
                document_type=document_type,
                extracted_text=extracted_text,
                extracted_fields=extracted_fields,
                confidence_score=0.8,  # Basic confidence score
                processing_time=processing_time
            )
            
        except Exception as e:
            print(f"Error processing document {file_path}: {e}")
            return DocumentExtractionResult(
                document_type=document_type,
                extracted_text="",
                extracted_fields=[],
                confidence_score=0.0,
                processing_time=time.time() - start_time
            )

